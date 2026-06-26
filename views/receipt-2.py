import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
import pandas as pd
from datetime import datetime
from pathlib import Path
import shutil
import config

from utils import (
    brand_green,
    brand_brown,
    logo_base64,
    get_baht_text,
    format_thai_date,
)
from receipt_db import (
    init_db, get_next_receipt_no, log_receipt, find_today_duplicate,
    PAYMENT_PREFIX, INSURANCE_PREFIX,
)
from receipt_storage import save_receipt_file

init_db()


# ==========================================
# 1. 🤖 ฟังก์ชันสกัดข้อมูลจากไฟล์ดิบ
# ==========================================
def extract_raw_receipt(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    data = {"name": "", "receipt_no": "", "address": "", "date": "", "items": []}

    def val(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    for r in range(1, 15):
        row_vals = [val(r, c) for c in range(1, 20)]
        row_str = " ".join(row_vals)

        if "ชื่อผู้ป่วย" in row_str and not data["name"]:
            for v in row_vals:
                if "ชื่อผู้ป่วย" in v:
                    data["name"] = v.replace("ชื่อผู้ป่วย", "").replace(":", "").strip()
        if "เลขที่" in row_str and "เสียภาษี" not in row_str and not data["receipt_no"]:
            for v in row_vals:
                if "เลขที่" in v:
                    data["receipt_no"] = v.replace("เลขที่", "").replace(":", "").strip()

        thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
                       "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        if not data["date"]:
            for v in row_vals:
                if any(m in v for m in thai_months) and "256" in v:
                    data["date"] = v

        # ดึง "ที่อยู่คนไข้" เฉพาะหลังจากเจอชื่อคนไข้แล้ว
        # (ที่อยู่คนไข้อยู่ใต้ชื่อ ส่วนที่อยู่ รพ. อยู่ในหัวกระดาษเหนือชื่อ จึงข้าม)
        if data["name"] and not data["address"]:
            for v in row_vals:
                if ("ต." in v and "อ." in v) or ("ถ." in v) or ("จ." in v):
                    if "โรงพยาบาล" not in v and "038-511" not in v and "เสียภาษี" not in v:
                        data["address"] = v

    for r in range(5, 50):
        cell_B, cell_A = val(r, 2), val(r, 1)
        item_no = cell_B if cell_B.isdigit() else cell_A
        if item_no.isdigit():
            item_name = ""
            for c in range(2, 10):
                v = val(r, c)
                if v and not v.isdigit() and v != "None":
                    item_name = v
                    break
            item_price = 0.0
            for c in range(20, 2, -1):
                v_raw = ws.cell(row=r, column=c).value
                if isinstance(v_raw, (int, float)):
                    item_price = float(v_raw)
                    break
                elif isinstance(v_raw, str):
                    try:
                        item_price = float(v_raw.replace(",", ""))
                        if item_price > 0:
                            break
                    except Exception:
                        pass
            if item_name:
                data["items"].append({"รายการ": item_name, "ราคา": item_price})
    return data


# ==========================================
# 2. ฟังก์ชันสร้างหน้าเว็บใบเสร็จ (HTML) สำหรับปริ้น
# ==========================================
def generate_receipt_html(data):
    items_html = ""
    for i, row in enumerate(data['items']):
        item_name = str(row.get('รายการ', '')).strip()
        item_price = row.get('ราคา', 0.0)
        if item_name:
            items_html += (
                f"<tr><td style='text-align:left;'>{i+1}</td>"
                f"<td style='text-align:left;'>{item_name}</td>"
                f"<td style='text-align:right;'>{item_price:,.2f}</td></tr>"
            )

    html = f"""<!DOCTYPE html><html lang="th"><head><meta charset="UTF-8"><style>
        * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
        html, body {{ height: 100%; }}
        body {{ font-family: 'Tahoma', sans-serif; color: #333; font-size: 14px; margin: 0; padding: 10px; box-sizing: border-box; }}
        .card {{ max-width: 800px; margin: 0 auto; padding: 20px; background-color: white; min-height: calc(100vh - 20px); display: flex; flex-direction: column; box-sizing: border-box; }}
        .header-container {{ position: relative; text-align: center; margin-bottom: 30px; }}
        .logo {{ position: absolute; left: 0; top: 0; width: 60px; }}
        .header-text {{ margin-left: 70px; }}
        .title {{ font-size: 18px; font-weight: bold; margin-bottom: 5px; }}
        .subtitle {{ font-size: 12px; margin-bottom: 2px; }}
        .info-row {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 13px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 15px; }}
        th, td {{ padding: 8px 5px; }}
        th {{ border-top: 1px solid #000; border-bottom: 1px solid #000; text-align: left; font-weight: bold; }}
        .total-row td {{ border-top: 1px solid #000; border-bottom: 1px solid #000; padding: 15px 5px; }}
        .footer-container {{ display: flex; justify-content: space-between; font-size: 11px; color: #555; margin-top: auto; padding-top: 30px; }}
        .print-btn {{ background-color: {brand_green}; color: white; border: none; padding: 10px; cursor: pointer; width: 100%; font-size: 16px; margin-bottom: 20px; font-weight:bold; }}
        @media print {{ body {{ padding: 0; }} .no-print {{ display: none !important; }} }}
    </style></head><body>
        <button class="no-print print-btn" onclick="window.print()">🖨️ ปริ้นใบเสร็จรับเงิน</button>
        <div class="card">
            <div class="header-container">
                <img src="data:image/png;base64,{logo_base64}" class="logo">
                <div class="header-text">
                    <div class="title">ใบเสร็จรับเงิน โรงพยาบาลโฮม ฉะเชิงเทรา</div>
                    <div class="subtitle">149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา โทร 038-511-123</div>
                    <div class="subtitle">เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
                </div>
            </div>
            <div class="info-row"><div>ชื่อผู้ป่วย <b>{data['name']}</b></div><div>เลขที่ {data['receipt_no']}</div></div>
            <div class="info-row"><div>{data['address']}</div><div>{data['date']}</div></div>
            <table>
                <tr><th style="width: 10%;">ลำดับที่</th><th style="width: 70%;">รายการ</th><th style="width: 20%; text-align:right;">ราคา</th></tr>
                {items_html}
                <tr class="total-row"><td colspan="2"><b>{data['baht_text']}</b></td><td style="text-align:right;"><b>{data['total']:,.2f}</b></td></tr>
            </table>
            <div class="info-row" style="margin-top: 20px;">
                <div><b>ชำระโดย : {data['payment_method']}</b></div>
                <div style="text-align: center; width: 300px;"><p>(........................................................)</p><p><b>{data['receiver']}</b><br>ผู้รับเงิน</p></div>
            </div>
            <div class="footer-container">
                <div><b>โรงพยาบาลโฮม ฉะเชิงเทรา</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา 24000<br>Tel. 038-511-123 | E-mail: Hospitalashome@gmail.com</div>
                <div style="text-align: right;"><b>บริษัท สื่อ การแพทย์ จำกัด</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง<br>อ.เมือง จ.ฉะเชิงเทรา 24000<br>เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
            </div>
        </div>
    </body></html>"""
    return base64.b64encode(html.encode('utf-8')).decode('utf-8')


# ==========================================
# 3. 📊 ฟังก์ชันสร้างไฟล์ Excel ใบเสร็จฉบับปรับปรุง
# ==========================================
def generate_receipt_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipt"
    ws.sheet_view.showGridLines = False

    FONT = "Tahoma"
    GREEN = "2C5E3B"
    GRAY = "555555"

    # ความกว้างคอลัมน์ (A=ลำดับ, B=รายการ, C=ราคา)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 22

    thin = Side(border_style="thin", color="000000")
    line_tb = Border(top=thin, bottom=thin)

    # ---------- โลโก้ (ฝังรูปมุมซ้ายบน) ----------
    try:
        logo_img = XLImage(BytesIO(base64.b64decode(logo_base64)))
        logo_img.width = 70
        logo_img.height = 70
        ws.add_image(logo_img, "A1")
    except Exception:
        pass

    # ---------- หัวกระดาษ (จัดกึ่งกลาง คร่อม A:C) ----------
    ws.merge_cells("A1:C1")
    ws["A1"] = "ใบเสร็จรับเงิน  โรงพยาบาลโฮม ฉะเชิงเทรา"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=GREEN)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    ws["A2"] = "149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา  โทร 038-511-123"
    ws["A2"].font = Font(name=FONT, size=10, color=GRAY)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:C3")
    ws["A3"] = "เลขประจำตัวผู้เสียภาษีอากร 0245563001367"
    ws["A3"].font = Font(name=FONT, size=10, color=GRAY)
    ws["A3"].alignment = Alignment(horizontal="center")

    # ---------- ข้อมูลผู้ป่วย ----------
    ws["A5"] = f"ชื่อผู้ป่วย : {data['name']}"
    ws["A5"].font = Font(name=FONT, size=11, bold=True)
    ws["C5"] = f"เลขที่ : {data['receipt_no']}"
    ws["C5"].font = Font(name=FONT, size=11, bold=True)
    ws["C5"].alignment = Alignment(horizontal="right")

    ws["A6"] = f"ที่อยู่ : {data['address']}"
    ws["A6"].font = Font(name=FONT, size=11)
    ws["C6"] = f"วันที่ : {data['date']}"
    ws["C6"].font = Font(name=FONT, size=11)
    ws["C6"].alignment = Alignment(horizontal="right")

    # ---------- หัวตาราง ----------
    head_row = 8
    ws[f"A{head_row}"] = "ลำดับที่"
    ws[f"B{head_row}"] = "รายการ"
    ws[f"C{head_row}"] = "ราคา (บาท)"
    for col in ("A", "B", "C"):
        c = ws[f"{col}{head_row}"]
        c.font = Font(name=FONT, size=11, bold=True)
        c.border = line_tb
        c.fill = PatternFill("solid", fgColor="EEF4EF")
    ws[f"A{head_row}"].alignment = Alignment(horizontal="center")
    ws[f"C{head_row}"].alignment = Alignment(horizontal="right")

    # ---------- รายการ ----------
    r = head_row + 1
    for i, item in enumerate(data["items"]):
        name = str(item.get("รายการ", "")).strip()
        if not name:
            continue
        ws[f"A{r}"] = i + 1
        ws[f"A{r}"].alignment = Alignment(horizontal="center")
        ws[f"B{r}"] = name
        ws[f"C{r}"] = item.get("ราคา", 0.0)
        ws[f"C{r}"].number_format = "#,##0.00"
        ws[f"C{r}"].alignment = Alignment(horizontal="right")
        for col in ("A", "B", "C"):
            ws[f"{col}{r}"].font = Font(name=FONT, size=11)
        r += 1

    # ---------- แถวรวมเงิน ----------
    ws[f"A{r}"] = data["baht_text"]
    ws[f"A{r}"].font = Font(name=FONT, size=11, bold=True)
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"C{r}"] = data["total"]
    ws[f"C{r}"].font = Font(name=FONT, size=12, bold=True, color=GREEN)
    ws[f"C{r}"].number_format = "#,##0.00"
    ws[f"C{r}"].alignment = Alignment(horizontal="right")
    for col in ("A", "B", "C"):
        ws[f"{col}{r}"].border = line_tb

    # ---------- ชำระเงิน + ผู้รับเงิน ----------
    pay_row = r + 2
    ws[f"A{pay_row}"] = f"ชำระโดย : {data['payment_method']}"
    ws[f"A{pay_row}"].font = Font(name=FONT, size=11, bold=True)

    ws[f"C{pay_row}"] = "(........................................)"
    ws[f"C{pay_row}"].alignment = Alignment(horizontal="center")
    ws[f"C{pay_row + 1}"] = data["receiver"]
    ws[f"C{pay_row + 1}"].font = Font(name=FONT, size=11, bold=True)
    ws[f"C{pay_row + 1}"].alignment = Alignment(horizontal="center")
    ws[f"C{pay_row + 2}"] = "ผู้รับเงิน"
    ws[f"C{pay_row + 2}"].font = Font(name=FONT, size=10, color=GRAY)
    ws[f"C{pay_row + 2}"].alignment = Alignment(horizontal="center")

    # ---------- footer ----------
    f_row = pay_row + 5
    ws.merge_cells(f"A{f_row}:B{f_row}")
    ws[f"A{f_row}"] = "โรงพยาบาลโฮม ฉะเชิงเทรา"
    ws[f"A{f_row}"].font = Font(name=FONT, size=9, bold=True, color=GRAY)
    ws.merge_cells(f"A{f_row + 1}:B{f_row + 1}")
    ws[f"A{f_row + 1}"] = "149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา 24000"
    ws[f"A{f_row + 1}"].font = Font(name=FONT, size=9, color=GRAY)
    ws.merge_cells(f"A{f_row + 2}:B{f_row + 2}")
    ws[f"A{f_row + 2}"] = "Tel. 038-511-123 | E-mail: Hospitalashome@gmail.com"
    ws[f"A{f_row + 2}"].font = Font(name=FONT, size=9, color=GRAY)

    ws[f"C{f_row}"] = "บริษัท สื่อ การแพทย์ จำกัด"
    ws[f"C{f_row}"].font = Font(name=FONT, size=9, bold=True, color=GRAY)
    ws[f"C{f_row}"].alignment = Alignment(horizontal="right")
    ws[f"C{f_row + 1}"] = "เลขประจำตัวผู้เสียภาษีอากร"
    ws[f"C{f_row + 1}"].font = Font(name=FONT, size=9, color=GRAY)
    ws[f"C{f_row + 1}"].alignment = Alignment(horizontal="right")
    ws[f"C{f_row + 2}"] = "0245563001367"
    ws[f"C{f_row + 2}"].font = Font(name=FONT, size=9, color=GRAY)
    ws[f"C{f_row + 2}"].alignment = Alignment(horizontal="right")

    # ---------- ตั้งค่าหน้ากระดาษสำหรับพิมพ์ ----------
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:C{f_row + 2}"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==========================================
# 4. หน้าจอ UI
# ==========================================
st.title("🧾 ระบบออกใบเสร็จรับเงิน รพ.โฮม (Auto-Fill System)")

# --- ส่วนที่ 1: เลือกไฟล์ใบเสร็จ ---
st.markdown("### 1️⃣ เลือกไฟล์ใบเสร็จ (ดึงข้อมูลเข้าระบบ)")

_inbox = getattr(config, "RAW_INBOX_DIR", "") or ""
_has_inbox = bool(_inbox) and Path(_inbox).is_dir()

source_mode = st.radio(
    "แหล่งไฟล์",
    ["📁 ดึงจากโฟลเดอร์อัตโนมัติ", "⬆️ อัปโหลดเอง"],
    horizontal=True,
    index=0 if _has_inbox else 1,
)

ex_data = {"name": "", "receipt_no": "", "address": "", "date": "ไม่ระบุวันที่", "items": []}
raw_path = None   # path ไฟล์ที่เลือกจากโฟลเดอร์ (ใช้ย้ายเข้า_done หลังออกใบเสร็จ)

def _load_from(src, file_id):
    """ตั้งค่า session เมื่อเป็นไฟล์ใหม่ + สกัดข้อมูล"""
    global ex_data
    if st.session_state.get("last_file_id") != file_id:
        st.session_state.last_file_id = file_id
        st.session_state.issued_no = None
        st.session_state.confirm_dup = False
        st.session_state.show_preview = False
    try:
        ex_data = extract_raw_receipt(src)
        st.success(
            f"✅ ดึงข้อมูลสำเร็จ! ชื่อ: {ex_data['name']} | "
            f"วันที่: {ex_data['date']} | รายการ: {len(ex_data['items'])} รายการ"
        )
    except Exception as e:
        st.error(f"❌ อ่านไฟล์ผิดพลาด: {e}")

if source_mode.startswith("📁"):
    if not _has_inbox:
        st.warning(
            f"⚠️ ยังไม่พบโฟลเดอร์ขาเข้า: `{_inbox or '(ยังไม่ได้ตั้งค่า)'}`\n\n"
            "ตั้งค่า **RAW_INBOX_DIR** ใน config.py ให้ชี้ไปยังโฟลเดอร์ที่ WinClinic export ไฟล์มา "
            "หรือเลือก '⬆️ อัปโหลดเอง'"
        )
    else:
        files = sorted(
            [p for p in Path(_inbox).glob("*.xls*") if p.is_file()],
            key=lambda p: p.stat().st_mtime, reverse=True
        )
        c_pick, c_btn = st.columns([4, 1])
        with c_btn:
            if st.button("🔄 รีเฟรช", width="stretch"):
                st.rerun()
        if not files:
            with c_pick:
                st.info("📭 ยังไม่มีไฟล์ในคิว — รอ WinClinic export ไฟล์มา แล้วกดรีเฟรช")
        else:
            with c_pick:
                _names = [f"{p.name}   ({datetime.fromtimestamp(p.stat().st_mtime):%H:%M:%S})"
                          for p in files]
                _idx = st.selectbox(
                    f"พบ {len(files)} ไฟล์ในคิว (ใหม่สุดอยู่บนสุด):",
                    range(len(files)), format_func=lambda i: _names[i]
                )
            raw_path = files[_idx]
            _load_from(str(raw_path), f"{raw_path.name}-{raw_path.stat().st_mtime}")
else:
    uploaded_receipt = st.file_uploader(
        "ลากไฟล์ Excel ใบเสร็จมาวางที่นี่...",
        type=["xlsx", "xls"], label_visibility="collapsed"
    )
    if uploaded_receipt:
        _load_from(uploaded_receipt, f"{uploaded_receipt.name}-{uploaded_receipt.size}")

# จำ path ไฟล์ดิบไว้ใน session (ใช้ย้ายเข้า _done หลังออกใบเสร็จ)
st.session_state.raw_path = str(raw_path) if raw_path else None

def _archive_raw():
    """ย้ายไฟล์ดิบที่ทำเสร็จไปโฟลเดอร์ย่อย done (กันประมวลผลซ้ำ)"""
    rp = st.session_state.get("raw_path")
    if not rp or not getattr(config, "MOVE_DONE_FILES", False):
        return
    try:
        src = Path(rp)
        if src.is_file():
            done_dir = src.parent / getattr(config, "DONE_FOLDER_NAME", "_done")
            done_dir.mkdir(parents=True, exist_ok=True)
            dest = done_dir / src.name
            n = 1
            while dest.exists():
                dest = done_dir / f"{src.stem} ({n}){src.suffix}"
                n += 1
            shutil.move(str(src), str(dest))
            st.session_state.raw_path = None
            st.session_state.last_file_id = None
    except Exception:
        pass  # ย้ายไม่ได้ก็ไม่ให้กระทบการออกใบเสร็จ


def _cleanup_old_done():
    """ลบไฟล์ในโฟลเดอร์ done ที่เก่ากว่า AUTO_DELETE_DAYS วัน (แนวทาง B)"""
    days = getattr(config, "AUTO_DELETE_DAYS", 0)
    inbox = getattr(config, "RAW_INBOX_DIR", "") or ""
    if not days or not inbox:
        return
    done_dir = Path(inbox) / getattr(config, "DONE_FOLDER_NAME", "_done")
    if not done_dir.is_dir():
        return
    cutoff = datetime.now().timestamp() - days * 86400
    for p in done_dir.glob("*"):
        try:
            if p.is_file() and p.stat().st_mtime < cutoff:
                p.unlink()
        except Exception:
            pass


# ล้างไฟล์เก่าครั้งเดียวต่อ session (ตอนเปิดหน้า)
if not st.session_state.get("_cleaned_done"):
    _cleanup_old_done()
    st.session_state._cleaned_done = True

st.divider()

# --- ส่วนที่ 2: ข้อมูลที่พนักงานกรอกเอง ---
st.markdown("### 2️⃣ ข้อมูลที่พนักงานต้องกรอกเอง")
col_emp1, col_emp2 = st.columns(2)
with col_emp1:
    receiver_name = st.text_input("👤 ชื่อผู้รับเงิน (พนักงาน)", value="นาย มรดก มาลี")
with col_emp2:
    cn_number = st.text_input("🆔 รหัสคนไข้ (HN / CN)", placeholder="(ไม่บังคับ)")

# เลือกประเภทการชำระ → กำหนดรหัสนำหน้าเลขใบเสร็จ
pay_type = st.radio(
    "💳 ประเภทการชำระ",
    ["เงินสด", "โอนจ่าย", "บัตรเครดิต", "ประกัน"],
    horizontal=True,
)

if pay_type == "ประกัน":
    col_ins1, col_ins2 = st.columns([1, 1])
    with col_ins1:
        insurer = st.selectbox(
            "🏢 เลือกบริษัทประกัน",
            list(INSURANCE_PREFIX.keys()),
            format_func=lambda x: f"{x}  ({INSURANCE_PREFIX[x]})",
        )
        prefix = INSURANCE_PREFIX[insurer]
        payment_label = f"ประกัน {insurer}"
    with col_ins2:
        st.markdown("**📋 ตารางรหัสบริษัทประกัน**")
        ins_df = pd.DataFrame(
            [{"บริษัทประกัน": k, "รหัสขึ้นต้น": v} for k, v in INSURANCE_PREFIX.items()]
        )
        st.dataframe(ins_df, hide_index=True, width='stretch')
else:
    prefix = PAYMENT_PREFIX[pay_type]
    payment_label = pay_type

_now = datetime.now()
_today_key = f"{(_now.year + 543) % 100:02d}{_now.month:02d}{_now.day:02d}"
st.info(f"รหัสนำหน้าเลขใบเสร็จ: **{prefix}**  →  ตัวอย่างเลขที่จะออก `{prefix}{_today_key}-01`")

st.divider()

# --- ส่วนที่ 3: ตารางรายการ + ดาวน์โหลด + พรีวิว ---
st.markdown("### 3️⃣ จัดการรายการและออกเอกสาร")
st.write("**ตารางรายการ (ดึงมาให้อัตโนมัติ เช็คความถูกต้องได้)**")
df_items = pd.DataFrame(ex_data["items"]) if ex_data["items"] else pd.DataFrame(
    [{"รายการ": "ค่าตรวจแพทย์ทั่วไป", "ราคา": 500.0}]
)

edited_df = st.data_editor(
    df_items,
    num_rows="dynamic",
    width='stretch',
    column_config={
        "รายการ": st.column_config.TextColumn("ชื่อรายการ", width="large", required=True),
        "ราคา": st.column_config.NumberColumn("ราคา (บาท)", min_value=0.0, format="%.2f"),
    }
)

total_fee = edited_df['ราคา'].sum()
baht_text_str = get_baht_text(total_fee)
st.info(f"**💰 รวมเงินทั้งสิ้น: {total_fee:,.2f} บาท** 👉 {baht_text_str}")

if "issued_no" not in st.session_state:
    st.session_state.issued_no = None
if "confirm_dup" not in st.session_state:
    st.session_state.confirm_dup = False

# ถ้าสลับโหมดการชำระ (prefix เปลี่ยน) ให้ล้างเลขเดิม กันเลขโหมดเก่าค้าง
if st.session_state.get("last_prefix") != prefix:
    st.session_state.last_prefix = prefix
    st.session_state.issued_no = None
    st.session_state.confirm_dup = False

# ถ้ายังไม่ออกเลข ใช้เลขจากไฟล์ Excel (ถ้ามี) หรือข้อความรอออกเลข
if st.session_state.issued_no:
    receipt_no_val = st.session_state.issued_no
else:
    receipt_no_val = ex_data["receipt_no"] if ex_data["receipt_no"] else "(ยังไม่ออกเลข)"

data_rec = {
    "name": ex_data["name"] if ex_data["name"] else "ไม่ระบุชื่อ",
    "hn": cn_number,
    "address": ex_data["address"],
    "receipt_no": receipt_no_val,
    "date": ex_data["date"] if ex_data["date"] != "ไม่ระบุวันที่" else format_thai_date(datetime.today()),
    "payment_method": payment_label,
    "receiver": receiver_name,
    "items": edited_df.to_dict('records'),
    "total": total_fee,
    "baht_text": baht_text_str,
}


def _issue_receipt():
    """ออกเลข + สร้างไฟล์ Excel (ฟอร์มสวย) + เซฟลงโฟลเดอร์ + log"""
    rno = get_next_receipt_no(prefix)            # จองเลขถัดไป (atomic)
    data_rec["receipt_no"] = rno
    # บันทึกไฟล์ Excel ฟอร์มสวย (โลโก้+หัวกระดาษ+ตาราง+footer)
    excel_io = generate_receipt_excel_file(data_rec)
    ok, info = save_receipt_file(rno, excel_io, data_rec["name"], ext="xlsx")
    log_receipt(rno, prefix, data_rec["name"], total_fee, payment_label, receiver_name)
    st.session_state.issued_no = rno
    st.session_state.last_receipt_data = dict(data_rec)   # เก็บไว้พรีวิว/ดาวน์โหลดหลังย้ายไฟล์
    st.session_state.show_preview = True                  # ให้พรีวิวเด้งอัตโนมัติ
    st.session_state.confirm_dup = False
    return ok, info, rno


def _resave_receipt(rno):
    """บันทึกซ้ำด้วยเลขเดิม (ไม่ออกเลขใหม่) — ไฟล์จะถูกเติม (1), (2) ต่อท้ายอัตโนมัติ"""
    data_rec["receipt_no"] = rno
    excel_io = generate_receipt_excel_file(data_rec)
    ok, info = save_receipt_file(rno, excel_io, data_rec["name"], ext="xlsx")
    st.session_state.issued_no = rno
    st.session_state.last_receipt_data = dict(data_rec)   # เก็บไว้พรีวิว/ดาวน์โหลดหลังย้ายไฟล์
    st.session_state.show_preview = True                  # ให้พรีวิวเด้งอัตโนมัติ
    st.session_state.confirm_dup = False
    return ok, info, rno


# --- ปุ่มออกเลขใบเสร็จ + บันทึกไฟล์ลงโฟลเดอร์แชร์ ---
col_btn1, col_btn2 = st.columns(2)
with col_btn1:
    if st.button("🧾 ออกเลขใบเสร็จ + บันทึกลงระบบ", type="primary", width='stretch'):
        if not ex_data["name"]:
            st.error("⚠️ ไม่พบชื่อผู้ป่วย กรุณาอัปโหลดไฟล์ใบเสร็จก่อนครับ")
        else:
            # ตรวจว่าคนไข้คนนี้ ยอดเท่ากัน ออกใบไปแล้ววันนี้หรือยัง
            dup_no = find_today_duplicate(data_rec["name"], total_fee, prefix)
            if dup_no and not st.session_state.confirm_dup:
                # เจอใบซ้ำ -> ขอให้กดยืนยันอีกครั้ง (ยังไม่บันทึก)
                st.session_state.confirm_dup = True
                st.session_state.dup_no = dup_no
            elif dup_no and st.session_state.confirm_dup:
                # ยืนยันบันทึกซ้ำ -> ใช้เลขเดิม ไม่ออกเลขใหม่ (ไฟล์เติม (1),(2) อัตโนมัติ)
                rno = st.session_state.get("dup_no", dup_no)
                ok, info, rno = _resave_receipt(rno)
                if ok:
                    _archive_raw()
                    st.success(f"✅ บันทึกซ้ำด้วยเลขเดิม **{rno}** แล้ว (ไฟล์เติมลำดับซ้ำให้อัตโนมัติ)\n\n📁 {info}")
                else:
                    st.warning(f"⚠️ บันทึกไฟล์ไม่สำเร็จ:\n\n{info}")
                st.rerun()
            else:
                # ไม่ซ้ำ -> ออกเลขใหม่ตามปกติ
                ok, info, rno = _issue_receipt()
                if ok:
                    _archive_raw()
                    st.success(f"✅ ออกเลข **{rno}** และบันทึกไฟล์แล้ว\n\n📁 {info}")
                else:
                    st.warning(f"⚠️ ออกเลข **{rno}** แล้ว แต่บันทึกไฟล์ลงโฟลเดอร์ไม่สำเร็จ:\n\n{info}")
                st.rerun()
with col_btn2:
    if st.button("🔄 เริ่มใบเสร็จใหม่ (ล้างเลข)", width='stretch'):
        st.session_state.issued_no = None
        st.session_state.confirm_dup = False
        st.session_state.last_receipt_data = None
        st.session_state.show_preview = False
        st.rerun()

# คำเตือนกรณีพบใบซ้ำ (ค้างไว้จนกว่าจะยืนยันหรือเริ่มใหม่)
if st.session_state.confirm_dup:
    st.warning(
        f"⚠️ **พบใบเสร็จซ้ำ!** คนไข้ **{data_rec['name']}** ยอด {total_fee:,.2f} บาท "
        f"ออกใบเสร็จไปแล้ววันนี้: **{st.session_state.get('dup_no', '-')}**\n\n"
        f"หากต้องการ**บันทึกซ้ำด้วยเลขเดิม** (ระบบจะเติม (1), (2) ต่อท้ายชื่อไฟล์ให้ ไม่ออกเลขใหม่) "
        f"กดปุ่ม \"🧾 ออกเลขใบเสร็จ + บันทึกลงระบบ\" อีกครั้งเพื่อยืนยัน "
        f"หรือกด \"🔄 เริ่มใบเสร็จใหม่\" เพื่อยกเลิก"
    )

if st.session_state.issued_no:
    st.success(f"เลขใบเสร็จปัจจุบัน: **{st.session_state.issued_no}**")

# ข้อมูลที่ใช้พรีวิว/ดาวน์โหลด: ถ้ามีไฟล์บนจอใช้ข้อมูลสด ถ้าไฟล์ถูกย้ายไปแล้วใช้ใบที่เพิ่งออก
_active_data = data_rec if ex_data.get("name") else (st.session_state.get("last_receipt_data") or data_rec)

# --- ดาวน์โหลด Excel ---
excel_receipt_data = generate_receipt_excel_file(_active_data)
safe_no = str(_active_data.get("receipt_no", receipt_no_val)).replace("(", "").replace(")", "").replace(" ", "")
st.download_button(
    label="📥 ดาวน์โหลดใบเสร็จรับเงินเป็นไฟล์ Excel",
    data=excel_receipt_data,
    file_name=f"Receipt_{safe_no}_{_active_data['name']}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width='stretch'
)

st.write("")

_do_preview = st.button("✨ สร้างหน้าพรีวิวสำหรับปริ้นใบเสร็จ", type="primary", width='stretch')
if _do_preview or st.session_state.get("show_preview"):
    if not _active_data.get("name") or _active_data["name"] == "ไม่ระบุชื่อ":
        if _do_preview:
            st.error("⚠️ ไม่พบข้อมูลใบเสร็จ กรุณาเลือกไฟล์ หรือออกเลขใบเสร็จก่อนครับ")
    else:
        iframe_code = (
            f'<iframe src="data:text/html;base64,{generate_receipt_html(_active_data)}" '
            f'width="100%" height="800" style="border:none;"></iframe>'
        )
        st.markdown(iframe_code, unsafe_allow_html=True)
