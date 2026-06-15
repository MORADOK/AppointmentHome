import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import base64
import pandas as pd
from datetime import datetime

from utils import (
    brand_green,
    brand_brown,
    logo_base64,
    get_baht_text,
    format_thai_date,
)
from receipt_db import (
    init_db, get_next_receipt_no, log_receipt,
    PAYMENT_PREFIX, INSURANCE_PREFIX,
)
from receipt_storage import save_receipt_file

init_db()


# ==========================================
# 1. 🤖 ฟังก์ชันสกัดข้อมูลจากไฟล์ดิบ
# ==========================================
def extract_raw_receipt(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    data = {"name": "", "receipt_no": "", "address": "", "date": "", "items": []}

    def val(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    for r in range(1, 15):
        row_vals = [val(r, c) for c in range(1, 20)]
        row_str = " ".join(row_vals)

        if "ชื่อผู้ป่วย" in row_str and not data["name"]:
            for v in row_vals:
                if "ชื่อผู้ป่วย" in v:
                    data["name"] = v.replace("ชื่อผู้ป่วย", "").replace(":", "").strip()
        if "เลขที่" in row_str and "เสียภาษี" not in row_str and not data["receipt_no"]:
            for v in row_vals:
                if "เลขที่" in v:
                    data["receipt_no"] = v.replace("เลขที่", "").replace(":", "").strip()

        thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
                       "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        if not data["date"]:
            for v in row_vals:
                if any(m in v for m in thai_months) and "256" in v:
                    data["date"] = v

        if not data["address"]:
            for v in row_vals:
                if ("ต." in v and "อ." in v) or ("ถ." in v) or ("จ." in v):
                    if "โรงพยาบาล" not in v:
                        data["address"] = v

    for r in range(5, 50):
        cell_B, cell_A = val(r, 2), val(r, 1)
        item_no = cell_B if cell_B.isdigit() else cell_A
        if item_no.isdigit():
            item_name = ""
            for c in range(2, 10):
                v = val(r, c)
                if v and not v.isdigit() and v != "None":
                    item_name = v
                    break
            item_price = 0.0
            for c in range(20, 2, -1):
                v_raw = ws.cell(row=r, column=c).value
                if isinstance(v_raw, (int, float)):
                    item_price = float(v_raw)
                    break
                elif isinstance(v_raw, str):
                    try:
                        item_price = float(v_raw.replace(",", ""))
                        if item_price > 0:
                            break
                    except Exception:
                        pass
            if item_name:
                data["items"].append({"รายการ": item_name, "ราคา": item_price})
    return data


# ==========================================
# 2. ฟังก์ชันสร้างหน้าเว็บใบเสร็จ (HTML) สำหรับปริ้น
# ==========================================
def generate_receipt_html(data):
    items_html = ""
    for i, row in enumerate(data['items']):
        item_name = str(row.get('รายการ', '')).strip()
        item_price = row.get('ราคา', 0.0)
        if item_name:
            items_html += (
                f"<tr><td style='text-align:left;'>{i+1}</td>"
                f"<td style='text-align:left;'>{item_name}</td>"
                f"<td style='text-align:right;'>{item_price:,.2f}</td></tr>"
            )

    html = f"""<!DOCTYPE html><html lang="th"><head><meta charset="UTF-8"><style>
        * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
        html, body {{ height: 100%; }}
        body {{ font-family: 'Tahoma', sans-serif; color: #333; font-size: 14px; margin: 0; padding: 10px; box-sizing: border-box; }}
        .card {{ max-width: 800px; margin: 0 auto; padding: 20px; background-color: white; min-height: calc(100vh - 20px); display: flex; flex-direction: column; box-sizing: border-box; }}
        .header-container {{ position: relative; text-align: center; margin-bottom: 30px; }}
        .logo {{ position: absolute; left: 0; top: 0; width: 60px; }}
        .header-text {{ margin-left: 70px; }}
        .title {{ font-size: 18px; font-weight: bold; margin-bottom: 5px; }}
        .subtitle {{ font-size: 12px; margin-bottom: 2px; }}
        .info-row {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 13px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 15px; }}
        th, td {{ padding: 8px 5px; }}
        th {{ border-top: 1px solid #000; border-bottom: 1px solid #000; text-align: left; font-weight: bold; }}
        .total-row td {{ border-top: 1px solid #000; border-bottom: 1px solid #000; padding: 15px 5px; }}
        .footer-container {{ display: flex; justify-content: space-between; font-size: 11px; color: #555; margin-top: auto; padding-top: 30px; }}
        .print-btn {{ background-color: {brand_green}; color: white; border: none; padding: 10px; cursor: pointer; width: 100%; font-size: 16px; margin-bottom: 20px; font-weight:bold; }}
        @media print {{ body {{ padding: 0; }} .no-print {{ display: none !important; }} }}
    </style></head><body>
        <button class="no-print print-btn" onclick="window.print()">🖨️ ปริ้นใบเสร็จรับเงิน</button>
        <div class="card">
            <div class="header-container">
                <img src="data:image/png;base64,{logo_base64}" class="logo">
                <div class="header-text">
                    <div class="title">ใบเสร็จรับเงิน โรงพยาบาลโฮม ฉะเชิงเทรา</div>
                    <div class="subtitle">149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา โทร 038-511-123</div>
                    <div class="subtitle">เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
                </div>
            </div>
            <div class="info-row"><div>ชื่อผู้ป่วย <b>{data['name']}</b></div><div>เลขที่ {data['receipt_no']}</div></div>
            <div class="info-row"><div>{data['address']}</div><div>{data['date']}</div></div>
            <table>
                <tr><th style="width: 10%;">ลำดับที่</th><th style="width: 70%;">รายการ</th><th style="width: 20%; text-align:right;">ราคา</th></tr>
                {items_html}
                <tr class="total-row"><td colspan="2"><b>{data['baht_text']}</b></td><td style="text-align:right;"><b>{data['total']:,.2f}</b></td></tr>
            </table>
            <div class="info-row" style="margin-top: 20px;">
                <div><b>ชำระโดย : {data['payment_method']}</b></div>
                <div style="text-align: center; width: 300px;"><p>(........................................................)</p><p><b>{data['receiver']}</b><br>ผู้รับเงิน</p></div>
            </div>
            <div class="footer-container">
                <div><b>โรงพยาบาลโฮม ฉะเชิงเทรา</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา 24000<br>Tel. 038-511-123 | E-mail: Hospitalashome@gmail.com</div>
                <div style="text-align: right;"><b>บริษัท สื่อ การแพทย์ จำกัด</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง<br>อ.เมือง จ.ฉะเชิงเทรา 24000<br>เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
            </div>
        </div>
    </body></html>"""
    return base64.b64encode(html.encode('utf-8')).decode('utf-8')


# ==========================================
# 3. 📊 ฟังก์ชันสร้างไฟล์ Excel ใบเสร็จฉบับปรับปรุง
# ==========================================
def generate_receipt_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipt"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

    ws['B1'] = "ใบเสร็จรับเงิน โรงพยาบาลโฮม ฉะเชิงเทรา"
    ws['B1'].font = Font(name='Tahoma', size=16, bold=True, color="2C5E3B")
    ws['B1'].alignment = Alignment(horizontal='center')

    ws['B2'] = "149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา โทร 038-511-123"
    ws['B2'].font = Font(name='Tahoma', size=10, color="555555")
    ws['B2'].alignment = Alignment(horizontal='center')

    ws['B3'] = "เลขประจำตัวผู้เสียภาษีอากร 0245563001367"
    ws['B3'].font = Font(name='Tahoma', size=10, color="555555")
    ws['B3'].alignment = Alignment(horizontal='center')

    ws['A5'] = f"ชื่อผู้ป่วย: {data['name']}"
    ws['A5'].font = Font(name='Tahoma', size=11, bold=True)
    ws['C5'] = f"เลขที่: {data['receipt_no']}"
    ws['C5'].font = Font(name='Tahoma', size=11)

    ws['A6'] = f"ที่อยู่: {data['address']}"
    ws['A6'].font = Font(name='Tahoma', size=11)
    ws['C6'] = f"วันที่: {data['date']}"
    ws['C6'].font = Font(name='Tahoma', size=11)

    ws['A8'] = "ลำดับที่"
    ws['B8'] = "รายการ"
    ws['C8'] = "ราคา"

    black_border_side = Side(border_style="thin", color="000000")
    header_border = Border(top=black_border_side, bottom=black_border_side)

    for col in ['A', 'B', 'C']:
        ws[f'{col}8'].font = Font(name='Tahoma', size=11, bold=True)
        ws[f'{col}8'].border = header_border
    ws['C8'].alignment = Alignment(horizontal='right')

    current_row = 9
    for i, item in enumerate(data['items']):
        if item.get('รายการ'):
            ws[f'A{current_row}'] = i + 1
            ws[f'B{current_row}'] = item['รายการ']
            ws[f'C{current_row}'] = item['ราคา']

            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = Font(name='Tahoma', size=11)
            ws[f'C{current_row}'].number_format = '#,##0.00'
            ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
            current_row += 1

    ws[f'A{current_row}'] = data['baht_text']
    ws[f'A{current_row}'].font = Font(name='Tahoma', size=11, bold=True)

    ws[f'C{current_row}'] = data['total']
    ws[f'C{current_row}'].font = Font(name='Tahoma', size=11, bold=True)
    ws[f'C{current_row}'].number_format = '#,##0.00'
    ws[f'C{current_row}'].alignment = Alignment(horizontal='right')

    total_border = Border(top=black_border_side, bottom=black_border_side)
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].border = total_border

    current_row += 2
    ws[f'A{current_row}'] = f"ชำระโดย : {data['payment_method']}"
    ws[f'A{current_row}'].font = Font(name='Tahoma', size=11, bold=True)

    ws[f'C{current_row}'] = "(........................................................)"
    ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    ws[f'C{current_row}'] = f"ผู้รับเงิน: {data['receiver']}"
    ws[f'C{current_row}'].font = Font(name='Tahoma', size=11, bold=True)
    ws[f'C{current_row}'].alignment = Alignment(horizontal='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==========================================
# 4. หน้าจอ UI
# ==========================================
st.title("🧾 ระบบออกใบเสร็จรับเงิน รพ.โฮม (Auto-Fill System)")

# --- ส่วนที่ 1: อัปโหลดไฟล์ ---
st.markdown("### 1️⃣ อัปโหลดไฟล์ใบเสร็จ Excel (เพื่อดึงข้อมูลเข้าระบบ)")
uploaded_receipt = st.file_uploader(
    "ลากไฟล์ Excel ใบเสร็จมาวางที่นี่...",
    type=["xlsx", "xls"],
    label_visibility="collapsed"
)

ex_data = {"name": "", "receipt_no": "", "address": "", "date": "ไม่ระบุวันที่", "items": []}
if uploaded_receipt:
    # ถ้าอัปโหลดไฟล์ใหม่ (ต่างจากเดิม) ให้ล้างเลขใบเสร็จเดิม กันเลขค้างข้ามคนไข้
    file_id = f"{uploaded_receipt.name}-{uploaded_receipt.size}"
    if st.session_state.get("last_file_id") != file_id:
        st.session_state.last_file_id = file_id
        st.session_state.issued_no = None
    try:
        ex_data = extract_raw_receipt(uploaded_receipt)
        st.success(
            f"✅ ดึงข้อมูลสำเร็จ! พบชื่อ: {ex_data['name']} | "
            f"วันที่: {ex_data['date']} | รายการ: {len(ex_data['items'])} รายการ"
        )
    except Exception as e:
        st.error(f"❌ อ่านไฟล์ผิดพลาด: {e}")

st.divider()

# --- ส่วนที่ 2: ข้อมูลที่พนักงานกรอกเอง ---
st.markdown("### 2️⃣ ข้อมูลที่พนักงานต้องกรอกเอง")
col_emp1, col_emp2 = st.columns(2)
with col_emp1:
    receiver_name = st.text_input("👤 ชื่อผู้รับเงิน (พนักงาน)", value="นาย มรดก มาลี")
with col_emp2:
    cn_number = st.text_input("🆔 รหัสคนไข้ (HN / CN)", placeholder="(ไม่บังคับ)")

# เลือกประเภทการชำระ → กำหนดรหัสนำหน้าเลขใบเสร็จ
pay_type = st.radio(
    "💳 ประเภทการชำระ",
    ["เงินสด", "โอนจ่าย", "บัตรเครดิต", "ประกัน"],
    horizontal=True,
)

if pay_type == "ประกัน":
    col_ins1, col_ins2 = st.columns([1, 1])
    with col_ins1:
        insurer = st.selectbox(
            "🏢 เลือกบริษัทประกัน",
            list(INSURANCE_PREFIX.keys()),
            format_func=lambda x: f"{x}  ({INSURANCE_PREFIX[x]})",
        )
        prefix = INSURANCE_PREFIX[insurer]
        payment_label = f"ประกัน {insurer}"
    with col_ins2:
        st.markdown("**📋 ตารางรหัสบริษัทประกัน**")
        ins_df = pd.DataFrame(
            [{"บริษัทประกัน": k, "รหัสขึ้นต้น": v} for k, v in INSURANCE_PREFIX.items()]
        )
        st.dataframe(ins_df, hide_index=True, width='stretch')
else:
    prefix = PAYMENT_PREFIX[pay_type]
    payment_label = pay_type

st.info(f"รหัสนำหน้าเลขใบเสร็จ: **{prefix}**  →  ตัวอย่างเลขที่จะออก `{prefix}690615-01`")

st.divider()

# --- ส่วนที่ 3: ตารางรายการ + ดาวน์โหลด + พรีวิว ---
st.markdown("### 3️⃣ จัดการรายการและออกเอกสาร")
st.write("**ตารางรายการ (ดึงมาให้อัตโนมัติ เช็คความถูกต้องได้)**")
df_items = pd.DataFrame(ex_data["items"]) if ex_data["items"] else pd.DataFrame(
    [{"รายการ": "ค่าตรวจแพทย์ทั่วไป", "ราคา": 500.0}]
)

edited_df = st.data_editor(
    df_items,
    num_rows="dynamic",
    width='stretch',
    column_config={
        "รายการ": st.column_config.TextColumn("ชื่อรายการ", width="large", required=True),
        "ราคา": st.column_config.NumberColumn("ราคา (บาท)", min_value=0.0, format="%.2f"),
    }
)

total_fee = edited_df['ราคา'].sum()
baht_text_str = get_baht_text(total_fee)
st.info(f"**💰 รวมเงินทั้งสิ้น: {total_fee:,.2f} บาท** 👉 {baht_text_str}")

if "issued_no" not in st.session_state:
    st.session_state.issued_no = None

# ถ้ายังไม่ออกเลข ใช้เลขจากไฟล์ Excel (ถ้ามี) หรือข้อความรอออกเลข
if st.session_state.issued_no:
    receipt_no_val = st.session_state.issued_no
else:
    receipt_no_val = ex_data["receipt_no"] if ex_data["receipt_no"] else "(ยังไม่ออกเลข)"

data_rec = {
    "name": ex_data["name"] if ex_data["name"] else "ไม่ระบุชื่อ",
    "hn": cn_number,
    "address": ex_data["address"],
    "receipt_no": receipt_no_val,
    "date": ex_data["date"] if ex_data["date"] != "ไม่ระบุวันที่" else format_thai_date(datetime.today()),
    "payment_method": payment_label,
    "receiver": receiver_name,
    "items": edited_df.to_dict('records'),
    "total": total_fee,
    "baht_text": baht_text_str,
}

# --- ปุ่มออกเลขใบเสร็จ + บันทึกไฟล์ลงโฟลเดอร์แชร์ ---
col_btn1, col_btn2 = st.columns(2)
with col_btn1:
    if st.button("🧾 ออกเลขใบเสร็จ + บันทึกลงระบบ", type="primary", width='stretch'):
        if not ex_data["name"]:
            st.error("⚠️ ไม่พบชื่อผู้ป่วย กรุณาอัปโหลดไฟล์ใบเสร็จก่อนครับ")
        else:
            rno = get_next_receipt_no(prefix)            # จองเลขถัดไป (atomic)
            data_rec["receipt_no"] = rno
            # บันทึกไฟล์รูปแบบเดียวกับหน้าพรีวิว (HTML พร้อมปริ้น โลโก้+footer ครบ)
            html_bytes = base64.b64decode(generate_receipt_html(data_rec))
            ok, info = save_receipt_file(rno, html_bytes, data_rec["name"], ext="html")
            log_receipt(rno, prefix, data_rec["name"], total_fee, payment_label, receiver_name)
            st.session_state.issued_no = rno
            if ok:
                st.success(f"✅ ออกเลข **{rno}** และบันทึกไฟล์แล้ว\n\n📁 {info}")
            else:
                st.warning(f"⚠️ ออกเลข **{rno}** แล้ว แต่บันทึกไฟล์ลงโฟลเดอร์ไม่สำเร็จ:\n\n{info}")
            st.rerun()
with col_btn2:
    if st.button("🔄 เริ่มใบเสร็จใหม่ (ล้างเลข)", width='stretch'):
        st.session_state.issued_no = None
        st.rerun()

if st.session_state.issued_no:
    st.success(f"เลขใบเสร็จปัจจุบัน: **{st.session_state.issued_no}**")

# --- ดาวน์โหลด Excel ---
excel_receipt_data = generate_receipt_excel_file(data_rec)
safe_no = receipt_no_val.replace("(", "").replace(")", "").replace(" ", "")
st.download_button(
    label="📥 ดาวน์โหลดใบเสร็จรับเงินเป็นไฟล์ Excel",
    data=excel_receipt_data,
    file_name=f"Receipt_{safe_no}_{data_rec['name']}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width='stretch'
)

st.write("")

if st.button("✨ สร้างหน้าพรีวิวสำหรับปริ้นใบเสร็จ", type="primary", width='stretch'):
    if not ex_data['name']:
        st.error("⚠️ ไม่พบชื่อผู้ป่วย กรุณาอัปโหลดไฟล์ใบเสร็จก่อนครับ")
    else:
        iframe_code = (
            f'<iframe src="data:text/html;base64,{generate_receipt_html(data_rec)}" '
            f'width="100%" height="800" style="border:none;"></iframe>'
        )
        st.markdown(iframe_code, unsafe_allow_html=True)