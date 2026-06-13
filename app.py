import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import base64
import pandas as pd
from datetime import datetime

# ==========================================
# 0. ตั้งค่าและซ่อนเมนู Streamlit (White-label App)
# ==========================================
st.set_page_config(page_title="ระบบจัดการ รพ.โฮม", page_icon="🏥", layout="wide")

hide_streamlit_style = """
<style>
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none !important;}
    #viewerBadge_container {display: none !important;}
    .viewerBadge_container {display: none !important;}
    [data-testid="viewerBadge"] {display: none !important;}
    .block-container { padding-top: 1rem; padding-bottom: 0rem; }
</style>
"""
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# ==========================================
# 1. ฟังก์ชันพื้นฐาน & ตัวแปลงเงินบาท
# ==========================================
def get_base64_image(image_path):
    try:
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode('utf-8')
    except Exception:
        return "" 

def get_baht_text(number):
    number = round(float(number), 2)
    if number == 0: return "ศูนย์บาทถ้วน"
    num_str = f"{number:,.2f}".replace(",", "")
    int_part, dec_part = num_str.split('.')
    digits = ["", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
    positions = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน", "ล้าน"]
    
    def read_num(n_str):
        res = ""
        length = len(n_str)
        for i, digit in enumerate(n_str):
            d = int(digit)
            pos = length - i - 1
            if d == 0: continue
            if pos == 0 and d == 1 and length > 1: res += "เอ็ด"
            elif pos == 1 and d == 1: res += "สิบ"
            elif pos == 1 and d == 2: res += "ยี่สิบ"
            else: res += digits[d] + positions[pos]
        return res
        
    text = read_num(int_part) + "บาท"
    if dec_part == "00": text += "ถ้วน"
    else: text += read_num(dec_part) + "สตางค์"
    return f"({text})"

def format_thai_date(date_obj):
    if not date_obj: return ""
    thai_months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
    return f"{date_obj.day} {thai_months[date_obj.month]} {date_obj.year + 543}"

# ==========================================
# 2. 🤖 ฟังก์ชันสกัดข้อมูลขั้นสูงจากไฟล์ดิบ
# ==========================================
def extract_raw_receipt(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    data = {"name": "", "receipt_no": "", "address": "", "date": "", "items": []}
    
    def val(r, c):
        v = ws.cell(row=r, column=c).value
        return str(v).strip() if v is not None else ""

    for r in range(1, 15):
        row_vals = [val(r, c) for c in range(1, 20)]
        row_str = " ".join(row_vals)
        
        if "ชื่อผู้ป่วย" in row_str and not data["name"]:
            for v in row_vals:
                if "ชื่อผู้ป่วย" in v: data["name"] = v.replace("ชื่อผู้ป่วย", "").replace(":", "").strip()
        if "เลขที่" in row_str and "เสียภาษี" not in row_str and not data["receipt_no"]:
            for v in row_vals:
                if "เลขที่" in v: data["receipt_no"] = v.replace("เลขที่", "").replace(":", "").strip()
        
        thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        if not data["date"]:
            for v in row_vals:
                if any(m in v for m in thai_months) and "256" in v: data["date"] = v
                    
        if not data["address"]:
            for v in row_vals:
                if ("ต." in v and "อ." in v) or ("ถ." in v) or ("จ." in v):
                    if "โรงพยาบาล" not in v: data["address"] = v

    for r in range(5, 50):
        cell_B, cell_A = val(r, 2), val(r, 1)
        item_no = cell_B if cell_B.isdigit() else cell_A
        if item_no.isdigit():
            item_name = ""
            for c in range(2, 10):
                v = val(r, c)
                if v and not v.isdigit() and v != "None":
                    item_name = v
                    break
            item_price = 0.0
            for c in range(20, 2, -1):
                v_raw = ws.cell(row=r, column=c).value
                if isinstance(v_raw, (int, float)):
                    item_price = float(v_raw)
                    break
                elif isinstance(v_raw, str):
                    try:
                        item_price = float(v_raw.replace(",", ""))
                        if item_price > 0: break
                    except: pass
            if item_name:
                data["items"].append({"รายการ": item_name, "ราคา": item_price})
    return data

# ==========================================
# 3. ฟังก์ชันสร้างหน้าเว็บสำหรับสั่งปริ้น (HTML)
# ==========================================
brand_green = "#2C5E3B"
brand_brown = "#8B5A2B"
logo_base64 = get_base64_image("logo.png")

def generate_receipt_html(data):
    items_html = ""
    for i, row in enumerate(data['items']):
        item_name = str(row.get('รายการ', '')).strip()
        item_price = row.get('ราคา', 0.0)
        if item_name:
            items_html += f"<tr><td style='text-align:left;'>{i+1}</td><td style='text-align:left;'>{item_name}</td><td style='text-align:right;'>{item_price:,.2f}</td></tr>"

    html = f"""<!DOCTYPE html><html lang="th"><head><meta charset="UTF-8"><style>
        * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
        body {{ font-family: 'Tahoma', sans-serif; color: #333; font-size: 14px; margin: 0; padding: 10px; }}
        .card {{ max-width: 800px; margin: 0 auto; padding: 20px; background-color: white; }}
        .header-container {{ position: relative; text-align: center; margin-bottom: 30px; }}
        .logo {{ position: absolute; left: 0; top: 0; width: 60px; }}
        .header-text {{ margin-left: 70px; }}
        .title {{ font-size: 18px; font-weight: bold; margin-bottom: 5px; }}
        .subtitle {{ font-size: 12px; margin-bottom: 2px; }}
        .info-row {{ display: flex; justify-content: space-between; margin-bottom: 5px; font-size: 13px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 15px; }}
        th, td {{ padding: 8px 5px; }}
        th {{ border-top: 1px solid #000; border-bottom: 1px solid #000; text-align: left; font-weight: bold; }}
        .total-row td {{ border-top: 1px solid #000; border-bottom: 1px solid #000; padding: 15px 5px; }}
        .footer-container {{ display: flex; justify-content: space-between; font-size: 11px; color: #555; margin-top: 50px; }}
        .print-btn {{ background-color: {brand_green}; color: white; border: none; padding: 10px; cursor: pointer; width: 100%; font-size: 16px; margin-bottom: 20px; font-weight:bold; }}
        @media print {{ body {{ padding: 0; }} .no-print {{ display: none !important; }} }}
    </style></head><body>
        <button class="no-print print-btn" onclick="window.print()">🖨️ ปริ้นใบเสร็จรับเงิน</button>
        <div class="card">
            <div class="header-container">
                <img src="data:image/png;base64,{logo_base64}" class="logo">
                <div class="header-text">
                    <div class="title">ใบเสร็จรับเงิน โรงพยาบาลโฮม ฉะเชิงเทรา</div>
                    <div class="subtitle">149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา โทร 038-511-123</div>
                    <div class="subtitle">เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
                </div>
            </div>
            <div class="info-row"><div>ชื่อผู้ป่วย <b>{data['name']}</b></div><div>เลขที่ {data['receipt_no']}</div></div>
            <div class="info-row"><div>{data['address']}</div><div>{data['date']}</div></div>
            <table>
                <tr><th style="width: 10%;">ลำดับที่</th><th style="width: 70%;">รายการ</th><th style="width: 20%; text-align:right;">ราคา</th></tr>
                {items_html}
                <tr class="total-row"><td colspan="2"><b>{data['baht_text']}</b></td><td style="text-align:right;"><b>{data['total']:,.2f}</b></td></tr>
            </table>
            <div class="info-row" style="margin-top: 20px;">
                <div><b>ชำระโดย : {data['payment_method']}</b></div>
                <div style="text-align: center; width: 300px;"><p>(........................................................)</p><p><b>{data['receiver']}</b><br>ผู้รับเงิน</p></div>
            </div>
            <div class="footer-container">
                <div><b>โรงพยาบาลโฮม ฉะเชิงเทรา</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา 24000<br>Tel. 038-511-123 | E-mail: Hospitalashome@gmail.com</div>
                <div style="text-align: right;"><b>บริษัท สื่อ การแพทย์ จำกัด</b><br>149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง<br>อ.เมือง จ.ฉะเชิงเทรา 24000<br>เลขประจำตัวผู้เสียภาษีอากร 0245563001367</div>
            </div>
        </div>
    </body></html>"""
    return base64.b64encode(html.encode('utf-8')).decode('utf-8')

def generate_appt_html(data):
    html = f"""<!DOCTYPE html><html lang="th"><head><meta charset="UTF-8"><style>
        * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
        body {{ font-family: 'Tahoma', sans-serif; color: #333; }}
        .card {{ max-width: 600px; margin: 0 auto; border: 1px solid #ddd; padding-bottom: 20px; background-color: white; }}
        .header {{ border-bottom: 2px solid {brand_green}; padding: 15px; display: flex; align-items: center; justify-content: center; margin-bottom: 15px; }}
        .logo {{ width: 60px; margin-right: 15px; }}
        .title {{ font-size: 20px; font-weight: bold; margin: 0; color: {brand_green}; }}
        .subtitle {{ font-size: 14px; margin: 0; color: #555; }}
        .content {{ padding: 0 20px; }}
        .row {{ display: flex; justify-content: space-between; margin-bottom: 10px; }}
        .print-btn {{ background-color: {brand_green}; color: white; border: none; padding: 10px; cursor: pointer; width: 100%; font-size: 16px; margin-bottom: 10px; }}
        @media print {{ .no-print {{ display: none !important; }} .card {{ border: none; }} }}
    </style></head><body>
        <button class="no-print print-btn" onclick="window.print()">🖨️ ปริ้นบัตรนัด</button>
        <div class="card">
            <div class="header"><img src="data:image/png;base64,{logo_base64}" class="logo">
            <div><p class="title">โรงพยาบาลโฮม ฉะเชิงเทรา</p><p class="subtitle">บัตรนัดหมาย | Appointment Card</p></div></div>
            <div class="content">
                <div class="row"><div><b>ชื่อ:</b> {data['name']}</div><div><b>HN:</b> {data['hn']}</div></div>
                <hr>
                <div class="row"><div style="color:{brand_green}; font-weight:bold;">วันที่นัด: {data['appt_date']}</div>
                <div style="color:{brand_green}; font-weight:bold;">เวลา: {data['appt_time']}</div></div>
                <div class="row"><div><b>แพทย์:</b> {data['doctor']}</div><div><b>รายการ:</b> {data['action']}</div></div>
                <p style="color:{brand_brown}; font-weight:bold;">📌 คำแนะนำ: <span style="color:#333; font-weight:normal;">{data['instruction']}</span></p>
                <p style="text-align:center; font-size:12px; margin-top:20px;">โทร 038-511-123 (กรุณานำยาเดิมมาด้วยทุกครั้ง)</p>
            </div>
        </div>
    </body></html>"""
    return base64.b64encode(html.encode('utf-8')).decode('utf-8')

# ==========================================
# 4. 📊 ฟังก์ชันสร้างไฟล์ Excel ใบเสร็จฉบับปรับปรุงอัตโนมัติ
# ==========================================
def generate_receipt_excel_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Receipt"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20

    ws['B1'] = "ใบเสร็จรับเงิน โรงพยาบาลโฮม ฉะเชิงเทรา"
    ws['B1'].font = Font(name='Tahoma', size=16, bold=True, color="2C5E3B")
    ws['B1'].alignment = Alignment(horizontal='center')
    
    ws['B2'] = "149/1 ถ.ฉะเชิงเทรา-บางปะกง ต.หน้าเมือง อ.เมือง จ.ฉะเชิงเทรา โทร 038-511-123"
    ws['B2'].font = Font(name='Tahoma', size=10, color="555555")
    ws['B2'].alignment = Alignment(horizontal='center')

    ws['B3'] = "เลขประจำตัวผู้เสียภาษีอากร 0245563001367"
    ws['B3'].font = Font(name='Tahoma', size=10, color="555555")
    ws['B3'].alignment = Alignment(horizontal='center')

    ws['A5'] = f"ชื่อผู้ป่วย: {data['name']}"
    ws['A5'].font = Font(name='Tahoma', size=11, bold=True)
    ws['C5'] = f"เลขที่: {data['receipt_no']}"
    ws['C5'].font = Font(name='Tahoma', size=11)
    
    ws['A6'] = f"ที่อยู่: {data['address']}"
    ws['A6'].font = Font(name='Tahoma', size=11)
    ws['C6'] = f"วันที่: {data['date']}"
    ws['C6'].font = Font(name='Tahoma', size=11)

    ws['A8'] = "ลำดับที่"
    ws['B8'] = "รายการ"
    ws['C8'] = "ราคา"
    
    black_border_side = Side(border_style="thin", color="000000")
    header_border = Border(top=black_border_side, bottom=black_border_side)
    
    for col in ['A', 'B', 'C']:
        ws[f'{col}8'].font = Font(name='Tahoma', size=11, bold=True)
        ws[f'{col}8'].border = header_border
    ws['C8'].alignment = Alignment(horizontal='right')

    current_row = 9
    for i, item in enumerate(data['items']):
        if item.get('รายการ'):
            ws[f'A{current_row}'] = i + 1
            ws[f'B{current_row}'] = item['รายการ']
            ws[f'C{current_row}'] = item['ราคา']
            
            for col in ['A', 'B', 'C']:
                ws[f'{col}{current_row}'].font = Font(name='Tahoma', size=11)
            ws[f'C{current_row}'].number_format = '#,##0.00'
            ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
            current_row += 1

    ws[f'A{current_row}'] = data['baht_text']
    ws[f'A{current_row}'].font = Font(name='Tahoma', size=11, bold=True)
    
    ws[f'C{current_row}'] = data['total']
    ws[f'C{current_row}'].font = Font(name='Tahoma', size=11, bold=True)
    ws[f'C{current_row}'].number_format = '#,##0.00'
    ws[f'C{current_row}'].alignment = Alignment(horizontal='right')

    total_border = Border(top=black_border_side, bottom=black_border_side)
    for col in ['A', 'B', 'C']:
        ws[f'{col}{current_row}'].border = total_border

    current_row += 2
    ws[f'A{current_row}'] = f"ชำระโดย : {data['payment_method']}"
    ws[f'A{current_row}'].font = Font(name='Tahoma', size=11, bold=True)

    ws[f'C{current_row}'] = "(........................................................)"
    ws[f'C{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    ws[f'C{current_row}'] = f"ผู้รับเงิน: {data['receiver']}"
    ws[f'C{current_row}'].font = Font(name='Tahoma', size=11, bold=True)
    ws[f'C{current_row}'].alignment = Alignment(horizontal='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# 5. หน้าจอ UI (Streamlit)
# ==========================================
st.title("🏥 ระบบให้บริการ รพ.โฮม (Auto-Fill System)")

# --- ส่วนที่ 1: อัปโหลดไฟล์กลาง ---
st.markdown("### 1️⃣ อัปโหลดไฟล์ใบเสร็จ Excel (เพื่อดึงข้อมูลเข้าระบบ)")
uploaded_receipt = st.file_uploader("ลากไฟล์ Excel ใบเสร็จมาวางที่นี่...", type=["xlsx", "xls"], label_visibility="collapsed")

ex_data = {"name": "", "receipt_no": "", "address": "", "date": "ไม่ระบุวันที่", "items": []}
if uploaded_receipt:
    try:
        ex_data = extract_raw_receipt(uploaded_receipt)
        st.success(f"✅ ดึงข้อมูลสำเร็จ! พบชื่อ: {ex_data['name']} | วันที่: {ex_data['date']} | รายการ: {len(ex_data['items'])} รายการ")
    except Exception as e:
        st.error(f"❌ อ่านไฟล์ผิดพลาด: {e}")

st.divider()

# --- ส่วนที่ 2: กรอกข้อมูลเพิ่มเติม ---
st.markdown("### 2️⃣ ข้อมูลที่พนักงานต้องกรอกเอง")
col_emp1, col_emp2, col_emp3 = st.columns(3)
with col_emp1: receiver_name = st.text_input("👤 ชื่อผู้รับเงิน (พนักงาน)", value="นาย มรดก มาลี")
with col_emp2: payment_method = st.selectbox("💳 วิธีการชำระเงิน", ["โอนจ่าย", "เงินสด", "บัตรเครดิต"])
with col_emp3: cn_number = st.text_input("🆔 รหัสคนไข้ (HN / CN)", placeholder="(เผื่อต้องการออกบัตรนัด)")

st.divider()

# --- ส่วนที่ 3: แท็บปริ้นและดาวน์โหลดเอกสาร ---
st.markdown("### 3️⃣ เลือกเอกสารที่ต้องการจัดการ")
tab1, tab2 = st.tabs(["🧾 จัดการใบเสร็จรับเงิน", "🏥 จัดการบัตรนัดหมาย"])

# --- แท็บ: ใบเสร็จรับเงิน ---
with tab1:
    st.write("**ตารางรายการ (ดึงมาให้อัตโนมัติ เช็คความถูกต้องได้)**")
    df_items = pd.DataFrame(ex_data["items"]) if ex_data["items"] else pd.DataFrame([{"รายการ": "ค่าตรวจแพทย์ทั่วไป", "ราคา": 500.0}])
    
    edited_df = st.data_editor(
        df_items,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "รายการ": st.column_config.TextColumn("ชื่อรายการ", width="large", required=True),
            "ราคา": st.column_config.NumberColumn("ราคา (บาท)", min_value=0.0, format="%.2f")
        }
    )
    
    total_fee = edited_df['ราคา'].sum()
    baht_text_str = get_baht_text(total_fee)
    st.info(f"**💰 รวมเงินทั้งสิ้น: {total_fee:,.2f} บาท** 👉 {baht_text_str}")

    receipt_no_val = ex_data["receipt_no"] if ex_data["receipt_no"] else "0000000000"

    data_rec = {
        "name": ex_data["name"] if ex_data["name"] else "ไม่ระบุชื่อ", 
        "hn": cn_number, 
        "address": ex_data["address"],
        "receipt_no": receipt_no_val,
        "date": ex_data["date"] if ex_data["date"] != "ไม่ระบุวันที่" else format_thai_date(datetime.today()),
        "payment_method": payment_method,
        "receiver": receiver_name,
        "items": edited_df.to_dict('records'),
        "total": total_fee,
        "baht_text": baht_text_str
    }

    excel_receipt_data = generate_receipt_excel_file(data_rec)
    st.download_button(
        label="📥 ดาวน์โหลดใบเสร็จรับเงินเป็นไฟล์ Excel",
        data=excel_receipt_data,
        file_name=f"Receipt_{receipt_no_val}_{data_rec['name']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    st.write("") 
    
    if st.button("✨ สร้างหน้าพรีวิวสำหรับปริ้นใบเสร็จ", type="primary", use_container_width=True):
        if not ex_data['name']:
            st.error("⚠️ ไม่พบชื่อผู้ป่วย กรุณาอัปโหลดไฟล์ใบเสร็จก่อนครับ")
        else:
            iframe_code = f'<iframe src="data:text/html;base64,{generate_receipt_html(data_rec)}" width="100%" height="800" style="border:none;"></iframe>'
            st.markdown(iframe_code, unsafe_allow_html=True)

# --- แท็บ: บัตรนัด ---
with tab2:
    appt_type = st.radio("ประเภทนัดหมาย", ["มาติดตามอาการ", "มาเจาะเลือด"], horizontal=True)
    
    # 🚨 นำฟังก์ชัน "งดน้ำและอาหาร" กลับมาตรงนี้ครับ
    default_action = "ตรวจติดตามอาการทั่วไป"
    default_instruction = "รับประทานยาและปฏิบัติตัวตามแพทย์สั่ง"

    if appt_type == "มาเจาะเลือด":
        fasting_option = st.selectbox("🥩 ต้องงดน้ำและงดอาหารหรือไม่?", ["ต้องงดน้ำและอาหาร", "ไม่ต้องงดน้ำและอาหาร"])
        if fasting_option == "ต้องงดน้ำและอาหาร":
            default_action = "เจาะเลือด FBS + Lipid ก่อนพบแพทย์"
            default_instruction = "งดน้ำ-งดอาหาร 6-8 ชั่วโมงก่อนตรวจ"
        else:
            default_action = "เจาะเลือดทั่วไป (ไม่ต้องงดอาหาร)"
            default_instruction = "ไม่ต้องงดน้ำและอาหาร สามารถรับประทานอาหารมาได้ตามปกติ"

    st.markdown("**📅 ระบบคำนวณวันนัดอัตโนมัติ**")
    col_c1, col_c2, col_c3 = st.columns([1, 1, 2])
    with col_c1:
        adv_num = st.number_input("จำนวนล่วงหน้า", min_value=0, value=0, step=1, help="ใส่ตัวเลข 0 หากต้องการใช้วันที่ปัจจุบัน")
    with col_c2:
        adv_unit = st.selectbox("หน่วย", ["วัน (Days)", "สัปดาห์ (Weeks)", "เดือน (Months)"])
    
    base_date = pd.Timestamp.today()
    if adv_num > 0:
        if "วัน" in adv_unit:
            calc_date = base_date + pd.DateOffset(days=adv_num)
        elif "สัปดาห์" in adv_unit:
            calc_date = base_date + pd.DateOffset(weeks=adv_num)
        else:
            calc_date = base_date + pd.DateOffset(months=adv_num)
    else:
        calc_date = base_date

    with col_c3:
        final_date = st.date_input("🗓️ ปฏิทิน (ระบบคำนวณให้อัตโนมัติ ปรับคลิกแก้ได้)", value=calc_date.date())
        appt_date_thai = format_thai_date(final_date)
        
    col_t1, col_t2 = st.columns(2)
    with col_t1: 
        appt_date_text = st.text_input("รูปแบบวันที่บนบัตร (แก้ไขข้อความได้)", value=appt_date_thai)
    with col_t2: 
        time_sel = st.selectbox("เวลานัด", ["08.00 - 10.00 น.", "10.00 - 12.00 น.", "13.00 - 15.00 น.", "15.00 - 16.30 น."])
        
    col_d1, col_d2 = st.columns(2)
    with col_d1: doc_appt = st.text_input("แพทย์ผู้ตรวจ", value="นพ.อภิสิทธิ์ สื่อประเสริฐสิทธิ์")
    with col_d2: act_appt = st.text_input("รายการตรวจ", value=default_action)
    ins_appt = st.text_input("คำแนะนำ", value=default_instruction)

    if st.button("✨ สร้างหน้าพรีวิวสำหรับปริ้นบัตรนัด", type="primary", use_container_width=True):
        if not ex_data['name']:
            st.error("⚠️ กรุณาอัปโหลดไฟล์ Excel ก่อนเพื่อดึงชื่อคนไข้ครับ")
        else:
            data_appt = {
                "name": ex_data["name"], 
                "hn": cn_number, 
                "type": appt_type, 
                "appt_date": appt_date_text, 
                "appt_time": time_sel, 
                "doctor": doc_appt, 
                "action": act_appt, 
                "instruction": ins_appt
            }
            st.markdown(f'<iframe src="data:text/html;base64,{generate_appt_html(data_appt)}" width="100%" height="500" style="border:none;"></iframe>', unsafe_allow_html=True)
